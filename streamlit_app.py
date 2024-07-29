import streamlit as st
import pandas as pd
import openai
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO

# Set up the OpenAI API key
api_key=st.secrets["OPENAI_API_KEY"]

# Title for the Streamlit app
st.title("Conversational Excel Sheet Assistant")

# File upload for source and target Excel files
source_file = st.file_uploader("Upload the source Excel file", type=["xlsx"])
target_file = st.file_uploader("Upload the target Excel file", type=["xlsx"])

if source_file and target_file:
    # Read the uploaded Excel files
    source_df = pd.read_excel(source_file)
    target_df = pd.read_excel(target_file)

    # Function to summarize dataframes
    def summarize_df(df):
        summary = {
            "columns": list(df.columns),
            "shape": df.shape,
            "dtypes": df.dtypes.to_dict(),
            "head": df.head().to_dict(),
            "tail": df.tail().to_dict(),
        }
        return summary

    source_summary = summarize_df(source_df)
    target_summary = summarize_df(target_df)

    # Function to find mismatches between source and target dataframes
    def find_mismatches(source_df, target_df):
        mismatches = []
        common_columns = set(source_df.columns).intersection(set(target_df.columns))
        for column in common_columns:
            source_col = source_df[column]
            target_col = target_df[column]
            for idx, (source_val, target_val) in enumerate(zip(source_col, target_col)):
                if pd.isnull(source_val) and pd.isnull(target_val):
                    continue
                if source_val != target_val:
                    mismatches.append({
                        "column": column,
                        "index": idx,
                        "source_value": source_val,
                        "target_value": target_val
                    })
        return mismatches

    mismatches = find_mismatches(source_df, target_df)

    # Output Excel file with mismatched rows highlighted in yellow
    def highlight_mismatches(source_file, source_df, target_df, mismatches):
        # Load the source and target files again with openpyxl to edit
        source_wb = load_workbook(source_file)
        source_ws = source_wb.active
        target_wb = load_workbook(target_file)
        target_ws = target_wb.active
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for mismatch in mismatches:
            row_idx = mismatch['index'] + 2  # Adding 2 because openpyxl is 1-indexed and header row is row 1
            # Highlight the entire row in source and target workbooks
            for col_idx in range(1, source_ws.max_column + 1):
                source_ws.cell(row=row_idx, column=col_idx).fill = fill
            for col_idx in range(1, target_ws.max_column + 1):
                target_ws.cell(row=row_idx, column=col_idx).fill = fill

        source_output = BytesIO()
        source_wb.save(source_output)
        source_output.seek(0)
        
        target_output = BytesIO()
        target_wb.save(target_output)
        target_output.seek(0)
        
        return source_output, target_output

    source_highlighted, target_highlighted = highlight_mismatches(source_file, source_df, target_df, mismatches)

    st.write("## Download Highlighted Mismatches")
    st.download_button("Download Source File with Highlighted Mismatches", data=source_highlighted, file_name="source_highlighted_mismatches.xlsx")
    st.download_button("Download Target File with Highlighted Mismatches", data=target_highlighted, file_name="target_highlighted_mismatches.xlsx")

    # Function to process user queries using OpenAI's GPT
    def answer_query(query, source_summary, target_summary, mismatches):
        if query.lower() in ["hi", "hello", "hey"]:
            return "Hello! How can I assist you with your Excel files today?"
        detailed_prompt = f"""
        You are an advanced validation chatbot designed to compare two Excel sheets: a source file and a target file. Analyze the summarized contents of both files and the mismatch report to answer the user's query with precise values and explanations.
        Source Data Summary:
        Columns: {source_summary['columns']}
        Shape: {source_summary['shape']}
        Data Types: {source_summary['dtypes']}
        Head: {source_summary['head']}
        Tail: {source_summary['tail']}
        Target Data Summary:
        Columns: {target_summary['columns']}
        Shape: {target_summary['shape']}
        Data Types: {target_summary['dtypes']}
        Head: {target_summary['head']}
        Tail: {target_summary['tail']}
        Mismatch Report: {mismatches}
        User query: {query}
        """
        try:
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are a highly sophisticated validation chatbot designed to perform rigorous comparisons between the contents of corresponding columns in a source file and a target file. Analyze the two Excel files in their entirety and respond to the user's query, providing the response with the correct and accurate values. Conduct a thorough, cell-by-cell comparison of each column's contents between the source and target files. Maintain unwavering diligence throughout the comparison process, as accuracy is of paramount importance. Ensure that your comparison logic can handle large datasets efficiently without compromising accuracy. Approach each comparison with the utmost care and precision, as your analysis may inform important business decisions and processes. Be aware of and account for potential discrepancies such as leading/trailing whitespace, case sensitivity, and numerical precision. Maintain a neutral, objective tone in your reporting, focusing on facts and data rather than assumptions."},
                    {"role": "user", "content": detailed_prompt}
                ],
                max_tokens=1500,
                temperature=0.1
            )
            answer = response.choices[0].message['content'].strip()
            return answer
        except Exception as e:
            return f"An error occurred: {e}"

    # Streamlit interface for user queries
    st.write("## Conversation History")
    # Initialize session state for user query and conversation history
    if "conversation" not in st.session_state:
        st.session_state.conversation = []

    # Function to submit user query and get a response
    def submit_query():
        user_query = st.session_state.user_query
        st.session_state.conversation.append({"role": "user", "content": user_query})
        response = answer_query(user_query, source_summary, target_summary, mismatches)
        st.session_state.conversation.append({"role": "assistant", "content": response})
        followup_question = generate_followup_question(st.session_state.conversation)
        st.session_state.conversation.append({"role": "assistant", "content": followup_question})
        st.session_state.user_query = ""  # Clear the input box after submitting

    def clear_input():
        st.session_state.user_query = ""  # Clear the input box

    # Function to generate follow-up questions based on the conversation
    def generate_followup_question(conversation):
        conversation_history = "\n".join([f"{msg['role']}: {msg['content']}" for msg in conversation])
        followup_prompt = f"""
        Based on the following conversation history, generate a relevant follow-up question for the user to keep the conversation going and gather more specific information:
        {conversation_history}
        """
        try:
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are an assistant generating follow-up questions based on the conversation history."},
                    {"role": "user", "content": followup_prompt}
                ],
                max_tokens=100,
                temperature=0.7
            )
            followup_question = response.choices[0].message['content'].strip()
            return followup_question
        except Exception as e:
            return f"An error occurred: {e}"

    # Display the conversation history
    for message in st.session_state.conversation:
        if message['role'] == 'user':
            st.write(f"**You:** {message['content']}")
        else:
            st.write(f"**Assistant:** {message['content']}")

    # Text input for user query at the bottom
    st.text_input("Enter your question or follow-up response about the Excel sheets:", key="user_query", on_change=submit_query)
    # Clear button to reset the input box
    st.button("Clear", on_click=clear_input)
    st.write('## Additional Information')

    # Display the dataframes
    st.write("## Source Data")
    st.dataframe(source_df)
    st.write("## Target Data")
    st.dataframe(target_df)

else:
    st.write("Please upload both source and target Excel files to proceed.")
